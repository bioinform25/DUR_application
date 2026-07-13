"""
HIRA 약제급여목록표 xlsx(latest.xlsx) -> data/drugs.json 변환.

원본 컬럼 (첫 행 헤더 기준):
연번 | 투여 | 분류 | 식약분류 | 주성분코드_동일제형 | 주성분코드 | 주성분갯수 |
주성분명 | 제품코드 | 제품명 | 업체명 | 규격 | 단위 | 상한금액표 금액 | 전일 | 비고

산출물 구조 (data/drugs.json)
------------------------------
{
  "updated": "2026-07-01",
  "source_title": "2026.07.01. 현재 약제급여목록및급여상한금액표",
  "product_count": N,
  "ingredient_count": M,
  "products": [
    {
      "product_code": "645302132",
      "product_name": "포크랄시럽(포수클로랄)_(9.5g/95mL)",
      "product_name_display": "포크랄시럽(포수클로랄)",
      "company": "한림제약(주)",
      "route": "내복",
      "spec": "95(1)",
      "unit": "mL/병",
      "price": 129,
      "drug_type": "전문",
      "ingredient_code": "130830ASY",
      "ingredient_code_form": "130830ASY",
      "ingredient_name": "chloral hydrate 9.5g(0.1g/mL)",
      "ingredient_name_display": "chloral hydrate"
    },
    ...
  ],
  "ingredients": {
    "130830ASY": {
      "name": "chloral hydrate",
      "product_codes": ["645302132", ...]
    },
    ...
  }
}
"""
import json
import re
import sys
from datetime import date
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from normalize import split_ingredient_keys  # noqa: E402

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
RAW_XLSX = DATA_DIR / "raw" / "latest.xlsx"
STATE_FILE = DATA_DIR / "state.json"
OUT_FILE = DATA_DIR / "drugs.json"

EXPECTED_HEADERS = [
    "연번", "투여", "분류", "식약분류", "주성분코드\n_동일제형", "주성분코드",
    "주성분\n갯수", "주성분명", "제품코드", "제품명", "업체명", "규격", "단위",
    "상한금액표 금액", "전일", "비고",
]

# 제품명 뒤에 붙는 "_(규격)" 표시부를 잘라 사람이 읽기 쉬운 이름을 만든다.
DISPLAY_NAME_RE = re.compile(r"^(.*?)_\([^()]*\)\s*$")
# 성분명에서 앞쪽 화학명만 남기고 뒤의 용량 표기를 잘라낸다. 예: "chloral hydrate   9.5g(0.1g/mL)" -> "chloral hydrate"
INGREDIENT_DISPLAY_RE = re.compile(r"^([A-Za-z0-9,\-\'\.\(\) ]+?)\s{2,}")


def strip_display_name(name: str) -> str:
    if not name:
        return name
    m = DISPLAY_NAME_RE.match(name)
    return m.group(1) if m else name


def strip_ingredient_display(name: str) -> str:
    if not name:
        return name
    name = name.strip()
    m = INGREDIENT_DISPLAY_RE.match(name)
    if m:
        candidate = m.group(1).strip()
        # 괄호가 짝이 맞지 않으면(예: "... (as doxepin" 처럼 중간에서 잘린 경우)
        # 잘라내지 않고 원본 전체를 사용한다.
        if candidate.count("(") == candidate.count(")"):
            return candidate
    return name


def load_state() -> dict:
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {}


def main() -> int:
    if not RAW_XLSX.exists():
        print(f"[parse_price_list] 원본 파일이 없습니다: {RAW_XLSX}")
        print("먼저 scripts/fetch_price_list.py 를 실행하세요.")
        return 1

    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header = [h.strip() if isinstance(h, str) else h for h in header]

    products = []
    ingredients = {}
    ingredient_key_index = {}

    for row in rows:
        if row is None or row[0] is None:
            continue
        record = dict(zip(header, row))

        product_code = record.get("제품코드")
        product_name = record.get("제품명")
        if not product_code or not product_name:
            continue

        ingredient_code = record.get("주성분코드") or ""
        ingredient_code_form = record.get("주성분코드\n_동일제형") or ingredient_code
        ingredient_name = record.get("주성분명") or ""
        price = record.get("상한금액표 금액")

        # 복합제 대응: "성분A 0.1g, 성분B 75mg" -> ["성분a", "성분b"] (DDI 매칭용 키)
        ingredient_keys = split_ingredient_keys(str(ingredient_name))

        product = {
            "product_code": str(product_code),
            "product_name": str(product_name),
            "product_name_display": strip_display_name(str(product_name)),
            "company": record.get("업체명") or "",
            "route": record.get("투여") or "",
            "spec": str(record.get("규격") or ""),
            "unit": record.get("단위") or "",
            "price": price,
            "drug_type": record.get("전일") or "",
            "ingredient_code": str(ingredient_code),
            "ingredient_code_form": str(ingredient_code_form),
            "ingredient_name": str(ingredient_name).strip(),
            "ingredient_name_display": strip_ingredient_display(str(ingredient_name)),
            "ingredient_keys": ingredient_keys,
        }
        products.append(product)

        if ingredient_code:
            entry = ingredients.setdefault(str(ingredient_code), {
                "name": product["ingredient_name_display"],
                "product_codes": [],
            })
            entry["product_codes"].append(product["product_code"])

        for key in ingredient_keys:
            ingredient_key_index.setdefault(key, [])
            if product["product_code"] not in ingredient_key_index[key]:
                ingredient_key_index[key].append(product["product_code"])

    state = load_state()
    out = {
        "updated": date.today().isoformat(),
        "source_title": state.get("title", ""),
        "product_count": len(products),
        "ingredient_count": len(ingredients),
        "products": products,
        "ingredients": ingredients,
        "ingredient_key_index": ingredient_key_index,
    }

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(
        json.dumps(out, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )

    size_mb = OUT_FILE.stat().st_size / (1024 * 1024)
    print(f"[parse_price_list] 제품 {len(products):,}개 / 성분 {len(ingredients):,}개")
    print(f"[parse_price_list] 저장 완료: {OUT_FILE} ({size_mb:.2f} MB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
